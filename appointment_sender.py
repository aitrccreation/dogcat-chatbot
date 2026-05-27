"""
Appointment Sender: Excel → LINE Flex Message
==============================================
รัน 18:00 ทุกวัน:
  สำหรับแต่ละแถวใน Send_Queue:
    - คำนวณ days_until = appt_date - today
    - days_until == 2 → ส่งครั้งเดียว → set sent_round_1_at
    - skip ถ้า status = Confirmed/Reschedule/Sent
    - skip ถ้า NoLine (ไม่มี line_user_id) — แต่บันทึก log แจ้ง admin

ใช้ Flex Message + Postback action สำหรับปุ่ม ยืนยัน/เลื่อน
"""
import io
import json
import os
import sys
from datetime import datetime, date, timedelta
from pathlib import Path

if sys.stdout and hasattr(sys.stdout, "buffer") and sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env", override=True)
except ImportError:
    pass

import requests
from openpyxl import load_workbook

XLSX = Path(__file__).parent / "appointments.xlsx"

# LINE OA token สำหรับส่ง reminder ให้ลูกค้า — ใช้ LINE_OA_TOKEN ก่อน, fallback LINE_TOKEN
LINE_TOKEN = os.environ.get(
    "LINE_OA_TOKEN",
    os.environ.get("LINE_TOKEN", "")
).strip()

# Admin notification → Wirote — ใช้ LOVELY_BOT_TOKEN (Wirote เป็นเพื่อน Lovely Bot, ไม่ใช่ LINE OA)
ADMIN_TOKEN = os.environ.get("LOVELY_BOT_TOKEN", "").strip() or LINE_TOKEN

ADMIN_LINE_ID = os.environ.get("LINE_TARGET_ID", "Ude09abe7b1f73ee901c047ccfe693dd8").strip()

THAI_MONTHS = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
               "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]


def thai_date(iso: str) -> str:
    """'2026-05-21' → '21 พ.ค. 2569'"""
    try:
        d = datetime.strptime(iso, "%Y-%m-%d").date()
        return f"{d.day} {THAI_MONTHS[d.month]} {d.year + 543}"
    except Exception:
        return iso


def thai_weekday(iso: str) -> str:
    try:
        d = datetime.strptime(iso, "%Y-%m-%d").date()
        return ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"][d.weekday()]
    except Exception:
        return ""


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ── Build Flex Message ──
def build_flex(qid: int, appt_date_iso: str, appt_time: str,
               pet_name: str, owner_name: str, vet: str, service: str,
               days_until: int, hn: str = "") -> dict:
    """LINE Flex Message bubble พร้อมปุ่มยืนยัน/เลื่อน"""
    appt_thai = thai_date(appt_date_iso)
    wd        = thai_weekday(appt_date_iso)
    time_str  = appt_time or "ตามเวลานัด"

    if days_until == 0:
        header_text = "🔔 นัดวันนี้!"
        header_color = "#DC2626"
    elif days_until == 1:
        header_text = "🔔 พรุ่งนี้มีนัดค่ะ"
        header_color = "#F59E0B"
    elif days_until == 2:
        header_text = "📅 แจ้งเตือนนัดล่วงหน้า 2 วัน"
        header_color = "#2563EB"
    else:
        header_text = "📅 แจ้งเตือนนัดล่วงหน้า"
        header_color = "#2563EB"

    bubble = {
        "type": "bubble",
        "size": "mega",
        "header": {
            "type":   "box",
            "layout": "vertical",
            "backgroundColor": header_color,
            "paddingAll": "16px",
            "contents": [
                {"type": "text", "text": header_text, "weight": "bold",
                 "size": "lg", "color": "#FFFFFF"},
            ],
        },
        "body": {
            "type":   "box",
            "layout": "vertical",
            "spacing": "md",
            "paddingAll": "16px",
            "contents": [
                {
                    "type": "box", "layout": "vertical", "spacing": "sm",
                    "contents": [
                        {"type": "text", "text": f"🐾 น้อง{pet_name or '-'}",
                         "weight": "bold", "size": "xl", "wrap": True},
                        {"type": "text", "text": f"เจ้าของ: {owner_name or '-'}",
                         "size": "sm", "color": "#6B7280"},
                    ],
                },
                {"type": "separator", "margin": "md"},
                {
                    "type": "box", "layout": "vertical", "spacing": "xs", "margin": "md",
                    "contents": [
                        {"type": "box", "layout": "baseline", "spacing": "sm",
                         "contents": [
                             {"type": "text", "text": "📅", "size": "sm", "flex": 0},
                             {"type": "text", "text": f"{appt_thai} ({wd})",
                              "size": "md", "weight": "bold", "wrap": True},
                         ]},
                        {"type": "box", "layout": "baseline", "spacing": "sm",
                         "contents": [
                             {"type": "text", "text": "⏰", "size": "sm", "flex": 0},
                             {"type": "text", "text": time_str, "size": "md", "wrap": True},
                         ]},
                        {"type": "box", "layout": "baseline", "spacing": "sm",
                         "contents": [
                             {"type": "text", "text": "🩺", "size": "sm", "flex": 0},
                             {"type": "text", "text": (service or "-")[:60],
                              "size": "sm", "wrap": True, "color": "#374151"},
                         ]},
                        {"type": "box", "layout": "baseline", "spacing": "sm",
                         "contents": [
                             {"type": "text", "text": "👨‍⚕️", "size": "sm", "flex": 0},
                             {"type": "text", "text": vet or "-", "size": "sm", "wrap": True},
                         ]},
                    ],
                },
            ],
        },
        "footer": {
            "type": "box", "layout": "vertical", "spacing": "sm",
            "contents": [
                {"type": "button", "style": "primary", "color": "#10B981",
                 "action": {"type": "postback",
                            "label": "✅ ยืนยันนัด",
                            "data": f"appt=confirm&qid={qid}&hn={hn}&pet={pet_name or '-'}&date={appt_thai}",
                            "displayText": "ยืนยันนัด"}},
                {"type": "button", "style": "secondary",
                 "action": {"type": "postback",
                            "label": "🕐 ขอเลื่อนนัด",
                            "data": f"appt=reschedule&qid={qid}&hn={hn}&pet={pet_name or '-'}&date={appt_thai}",
                            "displayText": "ขอเลื่อนนัด"}},
            ],
        },
    }

    alt_text = f"นัด {appt_thai} น้อง{pet_name}"
    return {"type": "flex", "altText": alt_text[:400], "contents": bubble}


# ── LINE push ──
def send_flex(line_user_id: str, flex_msg: dict, max_retries: int = 3) -> tuple[bool, str]:
    if not LINE_TOKEN:
        return False, "no LINE token"
    import time as _t
    url = "https://api.line.me/v2/bot/message/push"
    payload = {"to": line_user_id, "messages": [flex_msg]}
    last_err = ""
    for attempt in range(1, max_retries + 1):
        try:
            r = requests.post(
                url,
                headers={"Authorization": f"Bearer {LINE_TOKEN}",
                         "Content-Type": "application/json"},
                json=payload, timeout=15,
            )
            if r.status_code == 200:
                return True, ""
            last_err = f"HTTP {r.status_code} {r.text[:200]}"
            if 400 <= r.status_code < 500:
                return False, last_err  # don't retry 4xx
        except Exception as e:
            last_err = f"{type(e).__name__}: {e}"
        if attempt < max_retries:
            _t.sleep(attempt * 5)
    return False, last_err


# ── notify admin — ข้อความเดียวรวม pre-send + NoLine ──
def notify_admin_presend(ws, today: date) -> None:
    """สแกน Send_Queue และส่งสรุปครบในข้อความเดียวให้ Wirote:
      - T+2: จะส่ง LINE / ไม่มี LINE
      - วันนี้/พรุ่งนี้/3 วัน: ต้องโทรหาเอง
    (รวม notify_admin_nolist ไว้ในที่เดียว — ไม่ส่งซ้ำ)
    """
    if not LINE_TOKEN or not ADMIN_LINE_ID:
        return

    target_date = today + timedelta(days=2)
    target_iso  = target_date.strftime("%Y-%m-%d")
    target_thai = thai_date(target_iso)
    wd          = thai_weekday(target_iso)

    will_send:     list[str] = []   # T+2 + มี LINE + ยังไม่ส่ง
    no_line_t2:    list[str] = []   # T+2 + ไม่มี LINE
    no_line_call:  list[dict] = []  # วัน (3,1,0) + ไม่มี LINE → ต้องโทร

    SKIP_SERVICE_KEYWORDS = ("โทร",)
    CALL_DAYS = {3, 1, 0}

    for row_idx in range(2, ws.max_row + 1):
        qid           = ws.cell(row=row_idx, column=1).value
        appt_date_iso = ws.cell(row=row_idx, column=2).value
        hn            = ws.cell(row=row_idx, column=4).value or ""
        owner         = ws.cell(row=row_idx, column=5).value or ""
        pet           = ws.cell(row=row_idx, column=6).value or ""
        service       = ws.cell(row=row_idx, column=8).value or ""
        line_uid      = ws.cell(row=row_idx, column=9).value or ""
        status        = (ws.cell(row=row_idx, column=10).value or "").strip()
        r1_at         = ws.cell(row=row_idx, column=11).value

        if not qid or not appt_date_iso:
            continue
        if status in ("Confirmed", "Expired", "Cancel"):
            continue
        if any(k in str(service) for k in SKIP_SERVICE_KEYWORDS):
            continue

        try:
            appt_d = datetime.strptime(str(appt_date_iso), "%Y-%m-%d").date()
        except Exception:
            continue
        days_until = (appt_d - today).days

        is_no_line = not line_uid or status == "NoLine"
        label = f"น้อง{pet or '-'} ({owner or '-'}) HN {hn}"

        if str(appt_date_iso) == target_iso:          # T+2
            if is_no_line:
                no_line_t2.append(label)
            elif not r1_at:
                will_send.append(label)
        elif days_until in CALL_DAYS and is_no_line:  # วันใกล้ ต้องโทร
            no_line_call.append({
                "label":     label,
                "appt_date": str(appt_date_iso),
                "days_until": days_until,
            })

    # ไม่ส่งถ้าไม่มีรายการ
    if not will_send and not no_line_t2 and not no_line_call:
        return

    lines = [
        f"📋 สรุปนัดหมาย {target_thai} ({wd})",
        "━━━━━━━━━━━━━━━━━━━━━",
    ]

    # ส่วนที่ 1: จะส่ง LINE วันนี้ (T+2)
    if will_send:
        lines.append(f"✅ ส่ง LINE {len(will_send)} ราย:")
        for item in will_send[:20]:
            lines.append(f"  • {item}")
        if len(will_send) > 20:
            lines.append(f"  ... และอีก {len(will_send)-20} ราย")
    else:
        lines.append("✅ ส่ง LINE: ไม่มีรายการ")

    # ส่วนที่ 2: T+2 ไม่มี LINE
    if no_line_t2:
        lines.append(f"\n📵 ไม่มี LINE (นัดมะรืน) {len(no_line_t2)} ราย:")
        for item in no_line_t2[:10]:
            lines.append(f"  • {item}")
        if len(no_line_t2) > 10:
            lines.append(f"  ... และอีก {len(no_line_t2)-10} ราย")

    # ส่วนที่ 3: ต้องโทรวันนี้ (นัดใน 0-3 วัน ไม่มี LINE)
    if no_line_call:
        day_label = {0: "วันนี้", 1: "พรุ่งนี้", 3: "อีก 3 วัน"}
        lines.append(f"\n📞 ต้องโทรนัด {len(no_line_call)} ราย:")
        for r in no_line_call[:15]:
            dl = day_label.get(r["days_until"], f"อีก {r['days_until']} วัน")
            lines.append(f"  • [{dl}] {r['label']}")
        if len(no_line_call) > 15:
            lines.append(f"  ... และอีก {len(no_line_call)-15} ราย")

    text = "\n".join(lines)
    try:
        r = requests.post(
            "https://api.line.me/v2/bot/message/push",
            headers={"Authorization": f"Bearer {ADMIN_TOKEN}",
                     "Content-Type": "application/json"},
            json={"to": ADMIN_LINE_ID,
                  "messages": [{"type": "text", "text": text[:4900]}]},
            timeout=15,
        )
        if r.status_code == 200:
            print(f"  📋 Summary → Wirote: ส่ง {len(will_send)} | NoLine T+2 {len(no_line_t2)} | โทร {len(no_line_call)}")
        else:
            print(f"  [presend-notify] HTTP {r.status_code}: {r.text[:200]}")
    except Exception as e:
        print(f"  [presend-notify] error: {e}")


# ── notify_admin_nolist — ถูกรวมใน notify_admin_presend แล้ว (kept for compat) ──
def notify_admin_nolist(no_line_rows: list[dict]):
    pass  # merged into notify_admin_presend — no longer sends separate message


# ── MAIN ──
def main():
    args = set(sys.argv[1:])
    dry_run = "--dry-run" in args

    print("=" * 55)
    print("  📤 Appointment Sender — Excel → LINE")
    print(f"  Mode: {'DRY-RUN (ไม่ส่งจริง)' if dry_run else 'PRODUCTION'}")
    print("=" * 55)

    if not XLSX.exists():
        # บน Railway xlsx อาจหายตอน restart — สร้างใหม่จาก Google Sheet
        print(f"⚠️ {XLSX.name} missing — กำลังสร้างใหม่จาก Google Sheet...")
        try:
            import gsheet_db
            if gsheet_db.is_enabled():
                # restore Send_Queue จาก Google Sheet
                from openpyxl import Workbook
                ws_data = gsheet_db._get_queue()
                if ws_data:
                    wb_new = Workbook()
                    ws_new = wb_new.active
                    ws_new.title = "Send_Queue"
                    rows_data = ws_data.get_all_values()  # ทั้ง headers + rows
                    for r in rows_data:
                        ws_new.append(r)
                    wb_new.save(XLSX)
                    print(f"✅ restored {len(rows_data)-1} queue rows from Google Sheet")
                else:
                    print("❌ Google Sheet not configured — cannot send")
                    return
            else:
                print("❌ Google Sheet not configured — cannot send")
                return
        except Exception as e:
            print(f"❌ restore failed: {e}")
            return

    wb = load_workbook(XLSX)
    ws = wb["Send_Queue"]
    today = date.today()

    # ── ส่งสรุปให้ Wirote ก่อนส่งจริง ──
    if not dry_run:
        notify_admin_presend(ws, today)

    stats = {"sent_r1": 0, "sent_r2": 0, "sent_r3": 0, "skip": 0, "skip_service": 0, "noline": 0, "error": 0}
    noline_rows: list[dict] = []

    # บริการที่ไม่ต้องส่งเตือน (ตามที่ admin กำหนด)
    SKIP_SERVICE_KEYWORDS = ("โทร",)   # โทรสอบถาม/โทรถามอาการ — ไม่ใช่บริการจริง

    for row_idx in range(2, ws.max_row + 1):
        qid           = ws.cell(row=row_idx, column=1).value
        appt_date_iso = ws.cell(row=row_idx, column=2).value
        appt_time     = ws.cell(row=row_idx, column=3).value or ""
        hn            = ws.cell(row=row_idx, column=4).value or ""
        owner         = ws.cell(row=row_idx, column=5).value or ""
        pet           = ws.cell(row=row_idx, column=6).value or ""
        vet           = ws.cell(row=row_idx, column=7).value or ""
        service       = ws.cell(row=row_idx, column=8).value or ""
        line_uid      = ws.cell(row=row_idx, column=9).value or ""
        status        = (ws.cell(row=row_idx, column=10).value or "").strip()
        r1_at         = ws.cell(row=row_idx, column=11).value
        r2_at         = ws.cell(row=row_idx, column=12).value
        r3_at         = ws.cell(row=row_idx, column=13).value

        if not qid or not appt_date_iso:
            continue

        try:
            appt_d = datetime.strptime(str(appt_date_iso), "%Y-%m-%d").date()
        except Exception:
            continue

        days_until = (appt_d - today).days

        # Auto-expire rows ที่นัดผ่านแล้ว (days_until < 0)
        if days_until < 0 and status not in ("Sent", "Confirmed", "Reschedule", "Expired", "Cancel"):
            ws.cell(row=row_idx, column=10, value="Expired")
            stats["skip"] += 1
            continue

        # Skip terminal / expired states
        if status in ("Confirmed", "Expired", "Cancel"):
            stats["skip"] += 1
            continue

        # Skip บริการที่ไม่ต้องเตือน (โทรสอบถาม / โทรถาม)
        if any(k in str(service) for k in SKIP_SERVICE_KEYWORDS):
            stats["skip_service"] += 1
            print(f"  ⏭  qid={qid} HN={hn} น้อง{pet} — ข้าม (บริการ: {str(service)[:50]})")
            continue

        # Skip ที่ไม่ได้ลงทะเบียน — เก็บไว้แจ้ง admin
        if not line_uid or status == "NoLine":
            if days_until in (3, 1, 0):
                noline_rows.append({"hn": hn, "owner_name": owner, "pet_name": pet,
                                    "appt_date": str(appt_date_iso)})
                stats["noline"] += 1
            continue

        # Determine round — ส่งเฉพาะ T+2 ครั้งเดียว
        which_round = None
        if days_until == 2 and not r1_at:
            which_round = 1
        else:
            stats["skip"] += 1
            continue

        # Build + send (รองรับ line_uid = "uid1,uid2" สำหรับ 2 คนต่อ HN)
        flex = build_flex(qid, str(appt_date_iso), str(appt_time),
                          str(pet), str(owner), str(vet), str(service),
                          days_until, hn=str(hn))
        uid_list = [u.strip() for u in str(line_uid).split(",") if u.strip()]
        sent_count = 0
        for uid in uid_list:
            if dry_run:
                print(f"  [DRY] qid={qid} HN={hn} round={which_round} → {uid[:16]}...")
                sent_count += 1
            else:
                ok, err = send_flex(uid, flex)
                if ok:
                    sent_count += 1
                else:
                    stats["error"] += 1
                    print(f"  ❌ qid={qid} HN={hn} round={which_round} → {uid[:16]} FAIL: {err}")

        if sent_count > 0:
            # อัพ Excel ถ้าส่งได้อย่างน้อย 1 คน
            ts = now_iso()
            ws.cell(row=row_idx, column=10 + which_round, value=ts)
            if status == "Pending":
                ws.cell(row=row_idx, column=10, value="Sent")
            stats[f"sent_r{which_round}"] += 1
            print(f"  ✅ qid={qid} HN={hn} round={which_round} → {sent_count}/{len(uid_list)} recipient(s)")

    if not dry_run:
        wb.save(XLSX)
        # mirror Send_Queue ไป Google Sheet หลังส่ง (ให้ status เปลี่ยน)
        try:
            import gsheet_db
            if gsheet_db.is_enabled():
                from openpyxl import load_workbook as _lw
                wb_g = _lw(XLSX)
                ws_g = wb_g["Send_Queue"]
                rows_g = []
                for r in ws_g.iter_rows(min_row=2, values_only=True):
                    if r[0]:
                        rows_g.append([str(v) if v is not None else "" for v in r])
                gsheet_db.sync_send_queue(rows_g)
                print(f"   [gsheet] ✅ mirrored {len(rows_g)} queue rows after send")
        except Exception as e:
            print(f"   [gsheet] ⚠️ post-send mirror failed: {e}")

    print()
    print(f"📊 Summary: {stats}")
    if noline_rows:
        print(f"   {len(noline_rows)} รายที่ไม่มี LINE (รวมในสรุปก่อนส่งแล้ว)")
    return stats


if __name__ == "__main__":
    main()

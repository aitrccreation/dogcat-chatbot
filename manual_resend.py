"""
Manual Resend — qid=10006 HN 690224-1
ส่ง reminder ใหม่พร้อมข้อความขออภัยข้อมูลที่ผิดก่อนหน้า
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import os
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
load_dotenv()

import requests
import appointment_sender as sender
from openpyxl import load_workbook

XLSX = Path("appointments.xlsx")
LINE_TOKEN = os.environ.get("LINE_OA_TOKEN") or os.environ.get("LINE_TOKEN")

# ── ดึง qid=10006 จาก xlsx ──
wb = load_workbook(XLSX)
ws = wb["Send_Queue"]
target_row = None
for row_idx in range(2, ws.max_row + 1):
    if ws.cell(row=row_idx, column=1).value == 10006:
        target_row = row_idx
        break

if not target_row:
    print("❌ ไม่พบ qid=10006")
    sys.exit(1)

qid       = ws.cell(row=target_row, column=1).value
appt_iso  = ws.cell(row=target_row, column=2).value
appt_time = ws.cell(row=target_row, column=3).value or ""
hn        = ws.cell(row=target_row, column=4).value
owner     = ws.cell(row=target_row, column=5).value or ""
pet       = ws.cell(row=target_row, column=6).value or ""
vet       = ws.cell(row=target_row, column=7).value or ""
service   = ws.cell(row=target_row, column=8).value or ""
line_uid  = ws.cell(row=target_row, column=9).value

print(f"qid={qid} HN={hn} pet={pet} appt={appt_iso}")
print(f"  service: {service}")
print(f"  vet: {vet}")
print(f"  line_uid: {line_uid}")
print()

# ── ส่งข้อความอธิบายก่อน ──
uid_list = [u.strip() for u in str(line_uid).split(",") if u.strip()]

apology_text = (
    "🙏 ขออภัยค่ะ\n"
    "ข้อความนัดหมายก่อนหน้านี้ ข้อมูลในระบบยังไม่อัพเดต\n"
    "ขอแจ้งข้อมูลนัดที่ถูกต้องด้านล่างนี้นะคะ"
)

print("📨 กำลังส่ง...")
for uid in uid_list:
    # 1) ส่งข้อความขออภัย
    r1 = requests.post(
        "https://api.line.me/v2/bot/message/push",
        headers={"Authorization": f"Bearer {LINE_TOKEN}",
                 "Content-Type": "application/json"},
        json={"to": uid, "messages": [{"type": "text", "text": apology_text}]},
        timeout=15,
    )
    print(f"  [apology] uid={uid[:20]}... → {r1.status_code}")

    # 2) ส่ง Flex Message ที่ข้อมูลถูก
    flex = sender.build_flex(
        qid=qid,
        appt_date_iso=str(appt_iso),
        appt_time=str(appt_time),
        pet_name=str(pet),
        owner_name=str(owner),
        vet=str(vet),
        service=str(service),
        days_until=2,
        hn=str(hn),
    )
    ok, err = sender.send_flex(uid, flex)
    print(f"  [flex]    uid={uid[:20]}... → {'OK ✅' if ok else 'FAIL: ' + err}")

# ── อัพเดต status เป็น Sent ──
now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws.cell(row=target_row, column=10, value="Sent")
ws.cell(row=target_row, column=11, value=now)
ws.cell(row=target_row, column=15, value="manual_resend_correct_data")
wb.save(XLSX)
print(f"✅ อัพเดต qid={qid} status=Sent, sent_round_1_at={now}")

# ── Mirror ไป Google Sheet ──
try:
    import gsheet_db
    if gsheet_db.is_enabled():
        rows_g = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[0]:
                rows_g.append([str(v) if v is not None else "" for v in r])
        gsheet_db.sync_send_queue(rows_g)
        gsheet_db.log_event("Manual_Resend", hn=str(hn), line_user_id=line_uid,
                            detail=f"qid={qid} resent with corrected service/vet", result="OK")
        print(f"   [gsheet] ✅ mirrored {len(rows_g)} rows + logged")
except Exception as e:
    print(f"   [gsheet] ⚠️ {e}")

print()
print("🎉 เสร็จเรียบร้อย")

"""
Appointment Excel DB helper
============================
ทุก operation ที่อ่าน/เขียน appointments.xlsx ผ่านไฟล์นี้
- thread-safe (file lock)
- auto-create ถ้าไฟล์หาย
"""
import json
import re
import sys
import threading
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Bangkok TZ — บังคับ timestamp เป็น Asia/Bangkok เสมอ (Railway = UTC)
try:
    from zoneinfo import ZoneInfo
    _BKK_TZ = ZoneInfo("Asia/Bangkok")
except ImportError:
    try:
        import pytz
        _BKK_TZ = pytz.timezone("Asia/Bangkok")
    except ImportError:
        _BKK_TZ = None

XLSX_PATH = Path(__file__).parent / "appointments.xlsx"
_lock = threading.RLock()


# ───── ensure file + sheets exist ─────
def _ensure_file():
    if not XLSX_PATH.exists():
        # เรียก init script
        import subprocess
        subprocess.run([sys.executable, str(Path(__file__).parent / "init_appointments_xlsx.py")], check=True)
    return XLSX_PATH


def _load():
    _ensure_file()
    return load_workbook(XLSX_PATH)


def _save(wb):
    wb.save(XLSX_PATH)


def _now_iso() -> str:
    """คืนเวลาปัจจุบันใน Asia/Bangkok (UTC+7) เสมอ"""
    if _BKK_TZ is not None:
        return datetime.now(_BKK_TZ).strftime("%Y-%m-%d %H:%M:%S")
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ───── CUSTOMERS — userId ↔ HN mapping ─────
def find_customer_by_user_id(line_user_id: str) -> dict | None:
    with _lock:
        wb = _load()
        ws = wb["Customers"]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == line_user_id:
                return dict(zip(headers, row))
    return None


MAX_USERS_PER_HN = 2   # สูงสุด 2 LINE accounts ต่อ 1 HN


def find_customer_by_hn(hn: str) -> dict | None:
    """คืน customer แรกที่พบ (backward compat)"""
    result = find_customers_by_hn(hn)
    return result[0] if result else None


def find_customers_by_hn(hn: str) -> list[dict]:
    """คืน list ของ customers ทั้งหมดที่ register HN นี้ (max MAX_USERS_PER_HN)"""
    hn = (hn or "").strip()
    with _lock:
        wb = _load()
        ws = wb["Customers"]
        headers = [c.value for c in ws[1]]
        return [
            dict(zip(headers, row))
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[1] == hn and row[0]
        ]


def count_customers_by_hn(hn: str) -> int:
    """นับจำนวน LINE users ที่ register HN นี้"""
    return len(find_customers_by_hn(hn))


def get_line_uids_by_hn(hn: str) -> list[str]:
    """คืน list ของ line_user_id ทั้งหมดที่ผูกกับ HN นี้"""
    return [c["line_user_id"] for c in find_customers_by_hn(hn) if c.get("line_user_id")]


def _normalize_phone(phone) -> str:
    """คืนเบอร์เป็น string พร้อม 0 นำหน้า — gspread/xlsx ชอบแปลงเป็นตัวเลขแล้ว 0 หาย"""
    digits = "".join(ch for ch in str(phone or "") if ch.isdigit())
    if len(digits) == 9 and digits[0] in "2689":
        digits = "0" + digits
    return digits


def enrich_from_drx(hn: str, owner_name: str = "", pet_name: str = "",
                    pet_type: str = "", phone: str = "") -> tuple[str, str, str, str]:
    """เติมช่องที่ว่างจากข้อมูล DRX (SQLite cache → MySQL) — คืน (owner, pet, type, phone)
    เรียกทุกครั้งที่ลงทะเบียน เพื่อให้ column C/D/E/F ไม่มีช่องว่าง
    """
    phone = _normalize_phone(phone)
    if owner_name and pet_name and pet_type and phone:
        return owner_name, pet_name, pet_type, phone

    prof = None
    try:
        import opd_db
        prof = opd_db.get_pet_profile(hn)
    except Exception:
        pass

    if not prof or not (prof.get("pet_type") and prof.get("phone")):
        try:
            import drx_db
            r = drx_db.fetch_one(
                "SELECT p.petid, p.petname, p.pettype, "
                "       CONCAT(c.firstname,' ',c.lastname) AS owner, c.mobile_1, c.tel_1 "
                "FROM pet p LEFT JOIN customer c ON p.cuid = c.uid "
                "WHERE p.petid = %s LIMIT 1", (hn,)
            )
            if r:
                base = prof or {}
                prof = {
                    "pet_name":   base.get("pet_name")   or r.get("petname"),
                    "pet_type":   base.get("pet_type")   or r.get("pettype"),
                    "owner_name": base.get("owner_name") or r.get("owner"),
                    "phone":      base.get("phone")      or r.get("mobile_1") or r.get("tel_1"),
                }
        except Exception:
            pass

    if prof:
        # DRX เก็บคำนำหน้าไว้ใน full_name — "ไม่ระบุ" คือยังไม่ได้กรอก ไม่ใช่ชื่อจริง
        drx_owner = str(prof.get("owner_name") or "").strip()
        if drx_owner.startswith("ไม่ระบุ"):
            drx_owner = drx_owner[len("ไม่ระบุ"):].strip()
        owner_name = owner_name or drx_owner
        pet_name   = pet_name   or str(prof.get("pet_name")   or "").strip()
        pet_type   = pet_type   or str(prof.get("pet_type")   or "").strip()
        phone      = phone      or _normalize_phone(prof.get("phone"))

    return owner_name, pet_name, pet_type, phone


def register_customer(
    line_user_id: str,
    hn: str,
    owner_name: str = "",
    pet_name:   str = "",
    pet_type:   str = "",
    phone:      str = "",
    note:       str = "",
) -> dict | None:
    """เพิ่ม/อัพเดต customer mapping. Returns dict หรือ None ถ้า HN เต็มแล้ว
    เขียนทั้ง xlsx (local cache) และ Google Sheet (persistent) ถ้าตั้ง env vars แล้ว
    """
    owner_name, pet_name, pet_type, phone = enrich_from_drx(
        hn, owner_name, pet_name, pet_type, phone
    )
    # ── ลอง write ไป Google Sheet ก่อน (ถ้าตั้งไว้) ──
    try:
        import gsheet_db
        if gsheet_db.is_enabled():
            gs_result = gsheet_db.register_customer(
                line_user_id=line_user_id, hn=hn,
                owner_name=owner_name, pet_name=pet_name, pet_type=pet_type,
                phone=phone, note=note, max_per_hn=MAX_USERS_PER_HN,
            )
            if gs_result is None:
                # quota เต็ม — ไม่ต้องเขียน xlsx
                log_event("Register", hn, line_user_id, "HN quota full (gsheet)", "FAIL")
                return None
    except Exception as e:
        log_event("Register", hn, line_user_id, f"gsheet error: {e}", "WARN")
    with _lock:
        wb = _load()
        ws = wb["Customers"]
        now = _now_iso()

        # ── ตรวจ existing row โดยใช้ (uid, hn) เป็น key — ไม่ใช่ uid อย่างเดียว ──
        for row in ws.iter_rows(min_row=2):
            if row[0].value == line_user_id and str(row[1].value or "").strip() == hn.strip():
                # update existing
                if owner_name: row[2].value = owner_name
                if pet_name:   row[3].value = pet_name
                if pet_type:   row[4].value = pet_type
                if phone:      row[5].value = phone
                row[7].value = now
                if note:       row[8].value = note
                _save(wb)
                log_event("Register", hn, line_user_id, "updated", "OK")
                return _row_to_dict(ws, row[0].row)

        # ── คนใหม่ — ตรวจ quota ก่อน ──
        headers = [c.value for c in ws[1]]
        existing_count = sum(
            1 for row in ws.iter_rows(min_row=2, values_only=True)
            if row[1] == hn and row[0]
        )
        if existing_count >= MAX_USERS_PER_HN:
            log_event("Register", hn, line_user_id, f"HN full ({existing_count}/{MAX_USERS_PER_HN})", "FAIL")
            return None   # ← caller ต้องเช็ค None = ลงทะเบียนไม่ได้

        # ── append new ──
        ws.append([
            line_user_id, hn, owner_name, pet_name, pet_type, phone,
            now, now, note,
        ])
        _save(wb)
        log_event("Register", hn, line_user_id, f"new customer ({existing_count+1}/{MAX_USERS_PER_HN})", "OK")
        return {
            "line_user_id":  line_user_id, "hn": hn,
            "owner_name":    owner_name, "pet_name": pet_name,
            "pet_type":      pet_type, "phone": phone,
            "registered_at": now, "last_active": now, "note": note,
        }


def _row_to_dict(ws, row_idx: int) -> dict:
    headers = [c.value for c in ws[1]]
    return {h: ws.cell(row=row_idx, column=i+1).value for i, h in enumerate(headers)}


def get_all_customers() -> list[dict]:
    """คืน list ของ customer ทั้งหมดจาก Customers sheet"""
    with _lock:
        wb = _load()
        ws = wb["Customers"]
        headers = [c.value for c in ws[1]]
        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # มี line_user_id
                result.append(dict(zip(headers, row)))
        return result


# ───── LINE UID verification ─────
def verify_line_uid(line_user_id: str, token: str = "") -> bool | None:
    """ตรวจว่า LINE user ID นี้ยังใช้งานได้อยู่ (ไม่ block, ไม่ถูกลบ)
    Returns:
        True  = UID valid, bot สามารถส่งได้
        False = UID invalid / user block bot / account deleted
        None  = ไม่สามารถตรวจได้ (ไม่มี token หรือ network error)
    """
    import os
    if not token:
        token = (os.environ.get("LINE_OA_TOKEN") or os.environ.get("LINE_TOKEN", "")).strip()
    if not token or not line_user_id:
        return None
    try:
        import requests
        r = requests.get(
            f"https://api.line.me/v2/bot/profile/{line_user_id}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=8,
        )
        if r.status_code == 200:
            return True
        if r.status_code in (400, 404):
            return False   # UID ไม่มี หรือ block bot
        return None        # error อื่น (rate limit, server error) → ไม่ตัดสิน
    except Exception:
        return None


def update_customer_note(line_user_id: str, note: str):
    """อัพเดต note field ใน Customers sheet สำหรับ user นี้"""
    with _lock:
        wb = _load()
        ws = wb["Customers"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == line_user_id:
                # note อยู่ที่ column 9 (index 8)
                row[8].value = note
                row[7].value = _now_iso()   # last_active
                _save(wb)
                return True
    return False


def register_sibling_hn(
    line_user_id: str,
    hn: str,
    owner_name: str = "",
    pet_name:   str = "",
    pet_type:   str = "",
    phone:      str = "",
) -> dict | None:
    """ลงทะเบียน HN เพิ่มเติมสำหรับ LINE UID เดิม (เจ้าของมีหลายสัตว์)
    - ไม่ overwrite row เดิม — เพิ่ม row ใหม่เฉพาะคู่ (uid, hn) ที่ยังไม่มี
    - ตรวจ MAX_USERS_PER_HN เหมือนกัน
    - Returns dict ถ้าสำเร็จ, None ถ้า quota เต็มหรือมีอยู่แล้ว
    """
    hn = (hn or "").strip()
    if not hn or not line_user_id:
        return None

    owner_name, pet_name, pet_type, phone = enrich_from_drx(
        hn, owner_name, pet_name, pet_type, phone
    )

    # ── ลอง write ไป Google Sheet ก่อน (ถ้าตั้งไว้) ──
    try:
        import gsheet_db
        if gsheet_db.is_enabled():
            gsheet_db.register_customer(
                line_user_id=line_user_id, hn=hn,
                owner_name=owner_name, pet_name=pet_name,
                pet_type=pet_type, phone=phone,
                max_per_hn=MAX_USERS_PER_HN,
            )
    except Exception:
        pass

    with _lock:
        wb = _load()
        ws = wb["Customers"]
        now = _now_iso()

        # ตรวจว่า (uid, hn) คู่นี้มีอยู่แล้วหรือยัง
        for row in ws.iter_rows(min_row=2):
            if row[0].value == line_user_id and row[1].value == hn:
                # มีอยู่แล้ว — อัพข้อมูลถ้าขาด
                if owner_name and not row[2].value: row[2].value = owner_name
                if pet_name   and not row[3].value: row[3].value = pet_name
                if pet_type   and not row[4].value: row[4].value = pet_type
                if phone      and not row[5].value: row[5].value = phone
                _save(wb)
                return _row_to_dict(ws, row[0].row)

        # ตรวจ quota สำหรับ HN นี้
        existing_count = sum(
            1 for r in ws.iter_rows(min_row=2, values_only=True)
            if r[1] == hn and r[0]
        )
        if existing_count >= MAX_USERS_PER_HN:
            log_event("RegisterSibling", hn, line_user_id,
                      f"HN quota full ({existing_count}/{MAX_USERS_PER_HN})", "FAIL")
            return None

        # append row ใหม่
        ws.append([
            line_user_id, hn, owner_name, pet_name, pet_type, phone,
            now, now, "auto_sibling",
        ])
        _save(wb)
        log_event("RegisterSibling", hn, line_user_id,
                  f"sibling HN → {pet_name or hn}", "OK")
        return {
            "line_user_id": line_user_id, "hn": hn,
            "owner_name": owner_name, "pet_name": pet_name,
            "registered_at": now, "note": "auto_sibling",
        }


# ───── HN validator ─────
HN_PATTERN = re.compile(r"^\d{6}-\d+$")    # e.g. "690200-1"
HN_PATTERN_ALT = re.compile(r"^HN-?\d+(-\d+)?$", re.IGNORECASE)

def normalize_hn(text: str) -> str | None:
    """ดึง HN จากข้อความ user เช่น 'HN 690200-1', 'เลข 690200-1', '690200-1'"""
    if not text:
        return None
    # ลอง pattern ตรงๆ
    m = re.search(r"\b(\d{6,7}-\d+)\b", text)
    if m:
        return m.group(1)
    # ลอง HN- prefix
    m = re.search(r"HN[-\s]*(\d{6,7})(?:[-\s](\d+))?", text, re.IGNORECASE)
    if m:
        return f"{m.group(1)}-{m.group(2) or '1'}"
    return None


# ───── SEND_QUEUE updates ─────
def update_send_queue_status(queue_id: int, status: str, response: str = "") -> bool:
    """update status (Confirmed / Reschedule / Sent / ...)"""
    with _lock:
        wb = _load()
        ws = wb["Send_Queue"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == queue_id:
                row[9].value  = status       # status column J
                row[13].value = _now_iso()   # response_at column N
                row[14].value = response     # response column O
                _save(wb)
                return True
    return False


def find_queue_by_id(queue_id: int) -> dict | None:
    with _lock:
        wb = _load()
        ws = wb["Send_Queue"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == queue_id:
                headers = [c.value for c in ws[1]]
                return dict(zip(headers, row))
    return None


# ───── LOG ─────
def log_event(event: str, hn: str = "", line_user_id: str = "",
              detail: str = "", result: str = "OK"):
    """audit log row"""
    try:
        with _lock:
            wb = _load()
            ws = wb["Sent_Log"]
            ws.append([_now_iso(), event, hn, line_user_id, detail[:200], result])
            _save(wb)
    except Exception as e:
        # ห้าม raise — log fail ไม่ควร block flow
        print(f"[appointment_db] log_event error: {e}", file=sys.stderr)


# ───── DRX HN verification ─────
def verify_hn_with_drx(hn: str) -> dict | None:
    """ตรวจว่า HN นี้มีในระบบ DRX จริงไหม
    Returns: {hn, owner_name, pet_name, ...} หรือ None
    ใช้ข้อมูลจาก drx_data.json (อัพเดต daily) — fallback ถ้าไม่ได้
    """
    drx_file = Path(__file__).parent / "drx_data.json"
    if not drx_file.exists():
        return None
    try:
        data = json.loads(drx_file.read_text(encoding="utf-8"))
    except Exception:
        return None

    raw = data.get("_raw", {}) or {}

    # ── 1. ค้นจาก _raw.appointments (cell[11] = HN, cell[5] = "pet/owner") ──
    for c in raw.get("appointments", []):
        cells = c.get("cell", []) if isinstance(c, dict) else []
        if len(cells) > 11 and str(cells[11]).strip() == hn:
            pet_owner = str(cells[5]) if len(cells) > 5 else ""
            pet, owner = pet_owner.split("/", 1) if "/" in pet_owner else (pet_owner, "")
            return {
                "hn":         hn,
                "pet_name":   pet.strip(),
                "owner_name": owner.strip(),
                "phone":      str(cells[6]).strip() if len(cells) > 6 else "",
            }

    # ── 2. ค้นจาก _raw.cases (cell[2] = HN) ──
    for c in raw.get("cases", []):
        cells = c.get("cell", []) if isinstance(c, dict) else []
        if len(cells) >= 3 and str(cells[2]).strip() == hn:
            pet_owner = str(cells[1]) if len(cells) > 1 else ""
            pet, owner = pet_owner.split("/", 1) if "/" in pet_owner else (pet_owner, "")
            return {
                "hn":         hn,
                "pet_name":   pet.strip(),
                "owner_name": owner.strip(),
                "vet":        cells[4] if len(cells) > 4 else "",
            }

    # ── 3. ลอง mapped cases ──
    for c in data.get("cases", []):
        if c.get("hn") == hn or c.get("hn") == f"HN-{hn}":
            return {
                "hn":         hn,
                "pet_name":   c.get("pet", {}).get("name", ""),
                "owner_name": c.get("owner", ""),
                "vet":        c.get("vet", ""),
            }
    return None


# ───── for testing ─────
if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    print("=== Test normalize_hn ===")
    for t in ["690200-1", "HN 690200-1", "เลข HN690200", "ไม่มี hn"]:
        print(f"  {t!r} → {normalize_hn(t)}")
    print()
    print("=== Test verify_hn_with_drx ===")
    # use known HN from earlier captures
    for h in ["690257-1", "690261-1", "999999-1"]:
        info = verify_hn_with_drx(h)
        print(f"  {h}: {info}")

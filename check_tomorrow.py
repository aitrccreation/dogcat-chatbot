import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import load_workbook
from datetime import date, timedelta

wb = load_workbook("appointments.xlsx")
ws = wb["Send_Queue"]
headers = [c.value for c in ws[1]]

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0]:
        rows.append(dict(zip(headers, r)))

tomorrow = date.today() + timedelta(days=1)
print(f"พรุ่งนี้: {tomorrow} (23 พ.ค. 2569)")
print("=" * 60)

tmr = [r for r in rows if str(r.get("appt_date",""))[:10] == str(tomorrow)]
print(f"นัดหมายพรุ่งนี้ทั้งหมด: {len(tmr)} ราย")
print()

can_send = []
no_line  = []
already  = []

for r in tmr:
    lid    = str(r.get("line_user_id") or "").strip()
    status = str(r.get("status") or "").strip().lower()
    if status == "sent":
        already.append(r)
    elif lid and status not in ("skip","cancel","cancelled"):
        can_send.append(r)
    else:
        no_line.append(r)

print(f"✅ พร้อมส่ง 18:00 วันนี้: {len(can_send)} ราย")
for r in can_send:
    print(f"   qid={r['queue_id']}  HN={r['hn']}  น้อง{r['pet_name']}")
    print(f"   บริการ: {r.get('service','')}")
    print(f"   หมอ: {r.get('vet','')}  |  เวลา: {str(r.get('appt_time',''))[:5]}")
    print()

print(f"❌ ไม่มี LINE ID: {len(no_line)} ราย")
for r in no_line:
    print(f"   qid={r['queue_id']}  HN={r['hn']}  น้อง{r['pet_name']}  [{r.get('status')}]")

if already:
    print(f"\n📤 Sent แล้ว: {len(already)} ราย")
    for r in already:
        print(f"   qid={r['queue_id']}  HN={r['hn']}  น้อง{r['pet_name']}")

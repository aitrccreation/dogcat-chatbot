"""
Appointment Sync: DRX → Excel
==============================
รัน 06:00 ทุกวัน:
  1. ดึงนัดหมายจาก DRX (op=36 grid + op=903 calendar)
  2. กรองเฉพาะนัดในช่วง T+0 ถึง T+5 (ครอบคลุม 3-วันก่อน + 1-วันก่อน + วันนัด)
  3. เขียนลง Sheet "DRX_Appointments"
  4. รวมกับ Sheet "Manual" → Sheet "Send_Queue" (skip รายการที่ status เป็น Confirmed อยู่แล้ว)
  5. ลง Sent_Log

Usage:
    python appointment_sync.py             # ใช้ drx_data.json (cached)
    python appointment_sync.py --fetch     # ดึง DRX ใหม่ก่อน
"""
import io
import json
import re
import sys
import subprocess
from datetime import datetime, timedelta
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# Load .env so RAILWAY_BOT_URL / GOOGLE_SHEET_ID / INTERNAL_API_KEY are available
# when launched via Task Scheduler (where env vars are otherwise blank).
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent / ".env", override=False)
except Exception:
    pass

from openpyxl import load_workbook

XLSX = Path(__file__).parent / "appointments.xlsx"
DRX_JSON = Path(__file__).parent / "drx_data.json"


# ── helpers ──
def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


THAI_MONTH = {"ม.ค.": 1, "ก.พ.": 2, "มี.ค.": 3, "เม.ย.": 4, "พ.ค.": 5, "มิ.ย.": 6,
              "ก.ค.": 7, "ส.ค.": 8, "ก.ย.": 9, "ต.ค.": 10, "พ.ย.": 11, "ธ.ค.": 12}


def parse_thai_date(s: str) -> str | None:
    """'17 พ.ค. 2569' → '2026-05-17' (ISO)
    รองรับ Thai vowel marks เช่น 'มิ.ย.', 'มี.ค.' (ใช้ Unicode Thai range เต็ม)
    """
    if not s:
        return None
    # [฀-๿.] = Thai consonants + vowels + marks + dot
    m = re.match(r"(\d+)\s+([฀-๿.]+)\s+(\d{4})", s.strip())
    if not m:
        return None
    day, month_th, year = m.groups()
    mo = THAI_MONTH.get(month_th)
    if not mo:
        return None
    return f"{int(year) - 543:04d}-{mo:02d}-{int(day):02d}"


def fetch_drx_if_needed(force: bool = False):
    """รัน drx_bridge.run_fetch() ถ้า drx_data.json เก่ากว่า 24 ชม. หรือ force
    ใช้ import direct (ไม่ใช่ subprocess) เพื่อให้ทำงานได้บน Railway
    """
    if force or not DRX_JSON.exists():
        print("🔄 Fetching fresh DRX data...")
        try:
            import drx_bridge
            ok = drx_bridge.run_fetch()
            if not ok:
                print("[WARN] drx_bridge.run_fetch() returned False")
            else:
                print(f"   ✅ drx_data.json updated ({DRX_JSON.stat().st_size if DRX_JSON.exists() else 0} bytes)")
        except Exception as e:
            print(f"[ERROR] drx_bridge failed: {e}")


SKIP_DRX_STATUSES = ("มาตามนัด", "ยกเลิกนัดหมาย")


def load_drx_appointments() -> list[dict]:
    """อ่าน _raw.appointments (grid op=36) จาก drx_data.json
    Return: list of {appt_id, hn, owner_name, pet_name, vet, service, appt_date, appt_time, drx_status}
    """
    if not DRX_JSON.exists():
        print("❌ drx_data.json missing — run drx_bridge.py first")
        return []
    data = json.loads(DRX_JSON.read_text(encoding="utf-8"))

    raw = data.get("_raw", {}) or {}
    appts = []
    skipped_status = []

    # Source: _raw.appointments (grid op=36) — มี cell[16] = drx_status
    for r in raw.get("appointments", []):
        if not isinstance(r, dict):
            continue
        cells = r.get("cell", [])
        if len(cells) < 13:
            continue
        # cell layout:
        # [0] date   [5] "pet/owner"  [8] service category  [9] service detail
        # [11] HN    [12] hospital    [13] doctor            [16] status text
        appt_date = parse_thai_date(str(cells[0]))
        if not appt_date:
            continue
        drx_status = str(cells[16]).strip() if len(cells) > 16 and cells[16] else ""
        if drx_status in SKIP_DRX_STATUSES:
            skipped_status.append(f"{cells[5]} [{drx_status}]")
            continue
        pet_owner = str(cells[5]) if len(cells) > 5 else ""
        if "/" in pet_owner:
            pet, owner = pet_owner.split("/", 1)
        else:
            pet, owner = pet_owner, ""
        vet_doctor   = str(cells[13]) if len(cells) > 13 and cells[13] else ""
        vet_hospital = str(cells[12]) if len(cells) > 12 else ""
        vet_name = vet_doctor if vet_doctor and vet_doctor != vet_hospital else vet_hospital
        appts.append({
            "appt_id":    f"DRX-{r.get('id', '')}",
            "hn":         str(cells[11]) if len(cells) > 11 else "",
            "owner_name": owner.strip(),
            "pet_name":   pet.strip(),
            "vet":        vet_name,
            "service":    f"{cells[8]} — {cells[9]}" if len(cells) > 9 else str(cells[8] if len(cells) > 8 else ""),
            "appt_date":  appt_date,
            "appt_time":  "",
            "drx_status": drx_status,
            "source":     "drx",
        })

    if skipped_status:
        print(f"   ⏭  ข้าม {len(skipped_status)} รายการ (สถานะ DRX: มาตามนัด/ยกเลิก):")
        for s in skipped_status:
            print(f"      - {s}")

    return appts


def write_drx_sheet(appts: list[dict]):
    """เขียน DRX_Appointments sheet — overwrite ทั้ง sheet"""
    wb = load_workbook(XLSX)
    ws = wb["DRX_Appointments"]
    ws.delete_rows(2, ws.max_row)
    for a in appts:
        ws.append([
            a["appt_id"], a["hn"], a["owner_name"], a["pet_name"],
            a["vet"], a["service"], a["appt_date"], a["appt_time"],
            a["source"], now_iso(), a.get("drx_status", ""),
        ])
    wb.save(XLSX)
    print(f"✅ DRX_Appointments: {len(appts)} rows")


def collect_send_queue():
    """รวม DRX + Manual → Send_Queue ที่ตรงเงื่อนไข reminder window
    เงื่อนไข: appt_date เป็น T+1 หรือ T+2
      - T+2 = case ปกติ (build 13:00, ส่ง 18:00 ของวันเดียวกัน → ลูกค้าได้รับล่วงหน้า 2 วัน)
      - T+1 = catch-up นัดที่ DRX เพิ่งเพิ่มหลัง build T+2 ของเมื่อวาน
              (build 13:00 → ส่ง 18:00 → ลูกค้าได้รับล่วงหน้า 1 วัน)
    Skip: รายการที่อยู่ใน Send_Queue แล้วและ status=Confirmed (กันส่งซ้ำ)
    Skip: บริการที่เป็น "โทรสอบถาม/โทรถาม" (ไม่ใช่บริการจริง ไม่ต้องเตือน)
    """
    wb = load_workbook(XLSX)
    today = datetime.now().date()
    target_dates = {today + timedelta(days=1), today + timedelta(days=2)}

    def in_window(d: str) -> bool:
        try:
            dt = datetime.strptime(d, "%Y-%m-%d").date()
            return dt in target_dates  # T+1 หรือ T+2
        except Exception:
            return False

    SKIP_KEYWORDS = ("โทร",)   # โทรสอบถาม/โทรถามอาการ — ไม่ใช่บริการจริง ไม่ต้องเตือน

    def should_skip(service: str) -> bool:
        """True ถ้า service ตรงกับเงื่อนไขที่ไม่ต้องส่งเตือน"""
        s = (service or "")
        return any(k in s for k in SKIP_KEYWORDS)

    # Customer map: HN → comma-joined line_user_ids (สูงสุด MAX_USERS_PER_HN)
    # ข้าม UIDs ที่ถูก block / ลบบัญชี (note มี "LINE_BLOCKED")
    ws_c = wb["Customers"]
    hn_to_uids: dict[str, list[str]] = {}
    blocked_uid_count = 0
    for row in ws_c.iter_rows(min_row=2, values_only=True):
        if row[1] and row[0]:
            uid  = str(row[0]).strip()
            note = str(row[8] or "").strip() if len(row) > 8 else ""
            if "LINE_BLOCKED" in note:
                blocked_uid_count += 1
                continue   # UID นี้ส่งไม่ได้ — user block bot หรือลบบัญชีแล้ว
            hn_key = str(row[1]).strip()
            hn_to_uids.setdefault(hn_key, [])
            if uid not in hn_to_uids[hn_key]:
                hn_to_uids[hn_key].append(uid)
    if blocked_uid_count:
        print(f"   🚫 Skip {blocked_uid_count} LINE UIDs ที่ถูก block/ลบบัญชี")
    # รวมเป็น string "uid1,uid2" สำหรับใส่ใน queue
    hn_to_userid: dict[str, str] = {hn: ",".join(uids) for hn, uids in hn_to_uids.items()}

    # อ่าน existing queue → จำ confirmed
    ws_q = wb["Send_Queue"]
    existing_confirmed: set[tuple[str, str]] = set()  # (hn, appt_date)
    existing_keys: dict[tuple[str, str], int] = {}    # (hn, appt_date) → row_idx
    for row_idx, row in enumerate(ws_q.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]: continue
        key = (str(row[3] or ""), str(row[1] or ""))
        existing_keys[key] = row_idx
        if str(row[9] or "").lower() == "confirmed":
            existing_confirmed.add(key)

    # ── cleanup: mark expired rows (appt_date < today) as Expired ──
    expired_count = 0
    for row_idx, row in enumerate(ws_q.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]: continue
        appt_d_str = str(row[1] or "")[:10]
        status_val = str(row[9] or "").strip()
        r1_val = str(row[10] or "").strip()
        if status_val in ("Expired", "Sent", "Confirmed", "Reschedule", "Cancel"): continue
        try:
            appt_d = datetime.strptime(appt_d_str, "%Y-%m-%d").date()
            if appt_d < today:
                ws_q.cell(row=row_idx, column=10, value="Expired")
                expired_count += 1
        except Exception:
            pass
    if expired_count:
        print(f"   🗑️  Mark Expired: {expired_count} rows (นัดผ่านแล้ว)")

    # collect candidates
    skipped_by_service: list[dict] = []

    def collect_from_sheet(sheet_name: str, source: str):
        ws_s = wb[sheet_name]
        out = []
        for row in ws_s.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # appt_id
                continue
            appt_date = str(row[6] or "")
            if not in_window(appt_date):
                continue
            hn = str(row[1] or "").strip()
            key = (hn, appt_date)
            if key in existing_confirmed:
                continue
            service = str(row[5] or "")
            if should_skip(service):
                skipped_by_service.append({"hn": hn, "appt_date": appt_date,
                                            "pet": str(row[3] or ""),
                                            "service": service})
                continue
            out.append({
                "appt_id":    str(row[0]),
                "hn":         hn,
                "owner_name": str(row[2] or ""),
                "pet_name":   str(row[3] or ""),
                "vet":        str(row[4] or ""),
                "service":    service,
                "appt_date":  appt_date,
                "appt_time":  str(row[7] or ""),
                "source":     source,
            })
        return out

    drx_appts    = collect_from_sheet("DRX_Appointments", "drx")
    manual_appts = collect_from_sheet("Manual", "manual")
    raw_candidates = drx_appts + manual_appts

    if skipped_by_service:
        print(f"   ⏭  ข้าม {len(skipped_by_service)} รายการ (บริการที่ไม่ต้องเตือน):")
        for s in skipped_by_service:
            print(f"      - {s['appt_date']} HN={s['hn']} น้อง{s['pet']} | {s['service']}")

    # ── MERGE: รวม service หลายรายการต่อ HN+วันเดียวกันเป็น 1 entry ──
    # ป้องกันส่ง LINE หลายข้อความเมื่อ DRX มีหลาย service ในนัดเดียว
    merged: dict[tuple[str, str], dict] = {}
    for a in raw_candidates:
        key = (a["hn"], a["appt_date"])
        if key not in merged:
            merged[key] = dict(a)
        else:
            # รวม service โดยไม่ซ้ำ
            existing_svcs = [s.strip() for s in merged[key]["service"].split(" / ")]
            new_svc = a["service"].strip()
            if new_svc and new_svc not in existing_svcs:
                existing_svcs.append(new_svc)
                merged[key]["service"] = " / ".join(existing_svcs)
    candidates = list(merged.values())
    if len(raw_candidates) != len(candidates):
        print(f"   🔀  Merged: {len(raw_candidates)} DRX rows → {len(candidates)} queue entries (รวม service ซ้ำ HN+วัน)")

    # ── update Send_Queue ──
    next_qid = max([row[0] for row in ws_q.iter_rows(min_row=2, values_only=True) if row[0] and isinstance(row[0], int)] or [0]) + 1
    added = 0
    for a in candidates:
        key = (a["hn"], a["appt_date"])
        if key in existing_keys:
            # update existing row — รวม service ใหม่เข้าไปด้วยถ้ายังไม่มี
            row_idx = existing_keys[key]
            cur_svc = str(ws_q.cell(row=row_idx, column=8).value or "")
            new_svc = a["service"]
            if cur_svc != new_svc:
                # merge: เพิ่ม service ที่ยังไม่มีใน existing row
                cur_parts = [s.strip() for s in cur_svc.split(" / ") if s.strip()]
                for part in [s.strip() for s in new_svc.split(" / ") if s.strip()]:
                    if part not in cur_parts:
                        cur_parts.append(part)
                ws_q.cell(row=row_idx, column=8, value=" / ".join(cur_parts))
            ws_q.cell(row=row_idx, column=5, value=a["owner_name"])
            ws_q.cell(row=row_idx, column=6, value=a["pet_name"])
            ws_q.cell(row=row_idx, column=7, value=a["vet"])
            # update line_user_id ถ้าลูกค้าเพิ่งลงทะเบียน
            user_id = hn_to_userid.get(a["hn"], "")
            if user_id and not ws_q.cell(row=row_idx, column=9).value:
                ws_q.cell(row=row_idx, column=9, value=user_id)
                ws_q.cell(row=row_idx, column=10, value="Pending")
            continue
        # new row
        user_id = hn_to_userid.get(a["hn"], "")
        status = "Pending" if user_id else "NoLine"
        ws_q.append([
            next_qid, a["appt_date"], a["appt_time"], a["hn"],
            a["owner_name"], a["pet_name"], a["vet"], a["service"],
            user_id, status,
            "", "", "",     # sent_round_1/2/3_at
            "", "",          # response_at, response
            a["source"],
        ])
        existing_keys[key] = ws_q.max_row   # ✅ กันไม่ให้เพิ่มซ้ำในรอบเดียวกัน
        next_qid += 1
        added += 1

    wb.save(XLSX)
    return added, len(candidates)


def sync_drx_from_gsheet():
    """ดึง DRX_Appointments จาก Google Sheet → local xlsx
    ใช้บน Railway (PC sync DRX แล้ว upload Google Sheet)
    """
    try:
        import gsheet_db
        if not gsheet_db.is_enabled():
            print("   [drx-pull] gsheet not enabled — skip")
            return 0
        ws = gsheet_db._get_drx()
        if not ws:
            return 0
        records = ws.get_all_records()
        if not records:
            print("   [drx-pull] no DRX data ใน Google Sheet")
            return 0
        from openpyxl import load_workbook
        wb = load_workbook(XLSX)
        ws_xlsx = wb["DRX_Appointments"]
        # clear existing rows (keep header)
        if ws_xlsx.max_row > 1:
            ws_xlsx.delete_rows(2, ws_xlsx.max_row - 1)
        # write new rows
        for r in records:
            ws_xlsx.append([
                r.get("appt_id", ""),
                r.get("hn", ""),
                r.get("owner_name", ""),
                r.get("pet_name", ""),
                r.get("vet", ""),
                r.get("service", ""),
                r.get("appt_date", ""),
                r.get("appt_time", ""),
                r.get("source", ""),
                r.get("synced_at", ""),
            ])
        wb.save(XLSX)
        # แสดงวันที่ sync ล่าสุด (ป้องกันสับสนเมื่อ PC ไม่ได้รัน)
        synced_dates = [r.get("synced_at", "") for r in records if r.get("synced_at")]
        last_sync = max(synced_dates) if synced_dates else "ไม่ทราบ"
        today_str = datetime.now().strftime("%Y-%m-%d")
        stale_warn = " ⚠️ ข้อมูลเก่า (PC ไม่ได้ sync คืนนี้)" if last_sync[:10] < today_str else ""
        print(f"   [drx-pull] ✅ pulled {len(records)} DRX rows จาก Google Sheet (sync ล่าสุด: {last_sync[:16]}){stale_warn}")
        return len(records)
    except Exception as e:
        print(f"   [drx-pull] ⚠️ error: {e}")
        return 0


def sync_customers_from_gsheet():
    """ดึง customers จาก Google Sheet (persistent storage) → อัพ local Customers"""
    try:
        import gsheet_db
    except ImportError:
        return 0
    if not gsheet_db.is_enabled():
        return 0
    try:
        customers = gsheet_db.get_all_customers()
        if not customers:
            print("   [gsheet-sync] ไม่มี customers ใน Google Sheet")
            return 0
        import appointment_db as adb
        updated = 0
        for c in customers:
            uid = c.get("line_user_id")
            hn  = c.get("hn")
            if uid and hn:
                adb.register_customer(
                    line_user_id=uid, hn=hn,
                    owner_name=c.get("owner_name", ""),
                    pet_name=c.get("pet_name", ""),
                    phone=c.get("phone", ""),
                )
                updated += 1
        print(f"   [gsheet-sync] ✅ synced {updated} customers จาก Google Sheet")
        return updated
    except Exception as e:
        print(f"   [gsheet-sync] ⚠️ error: {e}")
        return 0


def sync_customers_from_railway():
    """ดึง customer registrations จาก Railway bot API → อัพ local Customers sheet
    ใช้ env var: RAILWAY_BOT_URL, INTERNAL_API_KEY
    (Fallback ถ้าไม่ได้ใช้ Google Sheet)
    """
    import os
    try:
        import requests as _req
    except ImportError:
        print("   [customer-sync] requests not installed — skip")
        return 0

    bot_url = os.environ.get("RAILWAY_BOT_URL", "").rstrip("/")
    api_key = os.environ.get("INTERNAL_API_KEY", "dogcatlovely_internal_2026")

    if not bot_url:
        print("   [customer-sync] RAILWAY_BOT_URL not set — skip")
        return 0

    try:
        r = _req.get(
            f"{bot_url}/api/customers",
            headers={"X-API-Key": api_key},
            timeout=15,
        )
        if r.status_code == 403:
            print("   [customer-sync] ❌ Unauthorized — check INTERNAL_API_KEY")
            return 0
        r.raise_for_status()
        data = r.json()
        customers = data.get("customers", [])
        if not customers:
            print("   [customer-sync] ไม่มีลูกค้าลงทะเบียนบน Railway")
            return 0

        import appointment_db as adb
        updated = 0
        for c in customers:
            uid = c.get("line_user_id")
            hn  = c.get("hn")
            if uid and hn:
                adb.register_customer(
                    line_user_id=uid,
                    hn=hn,
                    owner_name=c.get("owner_name") or "",
                    pet_name=c.get("pet_name") or "",
                    phone=c.get("phone") or "",
                )
                updated += 1
        print(f"   [customer-sync] ✅ synced {updated} customers จาก Railway")
        return updated
    except Exception as e:
        print(f"   [customer-sync] ⚠️ error: {e}")
        return 0


def _mirror_drx_to_gsheet():
    """Mirror DRX_Appointments จาก local xlsx ไป Google Sheet"""
    try:
        import gsheet_db
        if not gsheet_db.is_enabled():
            return
        from openpyxl import load_workbook
        wb_d = load_workbook(XLSX)
        ws_d = wb_d["DRX_Appointments"]
        rows_d = []
        for r in ws_d.iter_rows(min_row=2, values_only=True):
            if r[0]:
                rows_d.append([str(v) if v is not None else "" for v in r])
        if gsheet_db.sync_drx_appointments(rows_d):
            print(f"   [gsheet] ✅ mirrored {len(rows_d)} DRX rows to Google Sheet")
    except Exception as e:
        print(f"   [gsheet] ⚠️ sync_drx error: {e}")


def _mirror_queue_to_gsheet():
    """Mirror Send_Queue จาก local xlsx ไป Google Sheet"""
    try:
        import gsheet_db
        if not gsheet_db.is_enabled():
            return
        from openpyxl import load_workbook
        wb2 = load_workbook(XLSX)
        ws2 = wb2["Send_Queue"]
        rows = []
        for r in ws2.iter_rows(min_row=2, values_only=True):
            if r[0]:
                rows.append([str(v) if v is not None else "" for v in r])
        if gsheet_db.sync_send_queue(rows):
            print(f"   [gsheet] ✅ mirrored {len(rows)} Send_Queue rows to Google Sheet")
    except Exception as e:
        print(f"   [gsheet] ⚠️ sync_send_queue error: {e}")


def _hn_base(hn: str) -> str:
    """690131-2 → '690131'  (prefix ก่อน '-' = ตัวระบุเจ้าของ)"""
    return hn.split("-")[0] if hn and "-" in hn else hn


def _find_sibling_hns_from_drx(hn_base: str) -> list[dict]:
    """ค้นหา HNs ทั้งหมดที่มี prefix เดียวกัน (เจ้าของเดียวกัน) ใน DRX data
    เช่น hn_base='690131' → [{'hn':'690131-1','pet_name':'นมสด',...},
                               {'hn':'690131-2','pet_name':'กาแฟ',...}]
    """
    if not hn_base or not DRX_JSON.exists():
        return []
    try:
        data = json.loads(DRX_JSON.read_text(encoding="utf-8"))
        raw  = data.get("_raw", {}) or {}
        found: dict[str, dict] = {}   # hn → info

        # จาก _raw.appointments
        for r in raw.get("appointments", []):
            if not isinstance(r, dict): continue
            cells = r.get("cell", [])
            if len(cells) <= 11: continue
            hn = str(cells[11]).strip()
            if not hn or _hn_base(hn) != hn_base: continue
            pet_owner = str(cells[5]) if len(cells) > 5 else ""
            pet, owner = (pet_owner.split("/", 1) if "/" in pet_owner
                          else (pet_owner, ""))
            phone = str(cells[6]).strip() if len(cells) > 6 else ""
            if hn not in found:
                found[hn] = {"hn": hn, "pet_name": pet.strip(),
                              "owner_name": owner.strip(), "phone": phone}

        # จาก _raw.cases
        for r in raw.get("cases", []):
            if not isinstance(r, dict): continue
            cells = r.get("cell", [])
            if len(cells) < 3: continue
            hn = str(cells[2]).strip()
            if not hn or hn in found or _hn_base(hn) != hn_base: continue
            pet_owner = str(cells[1]) if len(cells) > 1 else ""
            pet, owner = (pet_owner.split("/", 1) if "/" in pet_owner
                          else (pet_owner, ""))
            found[hn] = {"hn": hn, "pet_name": pet.strip(),
                         "owner_name": owner.strip(), "phone": ""}

        return list(found.values())
    except Exception as e:
        print(f"   [siblings] drx scan error: {e}")
        return []


def _auto_register_sibling_hns(adb, all_customers: list[dict]) -> int:
    """สำหรับแต่ละ customer ที่มีชื่อแล้ว → ค้นหา HN พี่น้อง (prefix เดียวกัน)
    แล้ว register_sibling_hn() ให้อัตโนมัติ — LINE UID เดิมได้รับแจ้งเตือนทุกตัว

    หมายเหตุ: โหลด existing_pairs ล่าสุดจาก DB ทุกครั้ง (ไม่ใช้ snapshot)
    เพื่อป้องกัน duplicate เมื่อ gsheet sync เปลี่ยน HN หลักระหว่าง run
    """
    checked_bases: set[tuple[str, str]] = set()   # (uid, hn_base)
    new_count = 0

    for cust in all_customers:
        uid   = cust.get("line_user_id", "")
        hn    = cust.get("hn", "")
        owner = cust.get("owner_name", "").strip()
        # ข้ามถ้ายังไม่มีชื่อ หรือ HN ไม่มี "-"
        if not uid or not hn or not owner or "-" not in hn:
            continue
        base = _hn_base(hn)
        key  = (uid, base)
        if key in checked_bases:
            continue
        checked_bases.add(key)

        # ── โหลด existing pairs ล่าสุดจาก DB ทุกครั้ง (ไม่ใช้ snapshot) ──
        fresh_customers = adb.get_all_customers()
        existing_pairs = {
            (c["line_user_id"], c["hn"])
            for c in fresh_customers
            if c.get("line_user_id") and c.get("hn")
        }

        siblings = _find_sibling_hns_from_drx(base)
        for sib in siblings:
            sib_hn = sib["hn"]
            if (uid, sib_hn) in existing_pairs:
                continue   # มีอยู่แล้ว (ทั้ง primary และ auto_sibling)
            result = adb.register_sibling_hn(
                line_user_id=uid,
                hn=sib_hn,
                owner_name=sib.get("owner_name", owner),
                pet_name=sib.get("pet_name", ""),
                phone=sib.get("phone", ""),
            )
            if result:
                print(f"      🔗 Sibling HN {sib_hn} ({sib.get('pet_name','?')}) → UID ...{uid[-8:]}")
                new_count += 1
    return new_count


def backfill_customer_names():
    """หลัง DRX sync — ค้นหา customers ที่ลงทะเบียน HN ไว้แต่ยังไม่มีชื่อ
    แล้ว backfill owner_name / pet_name จาก drx_data.json อัตโนมัติ
    จากนั้น auto-register HN พี่น้อง (เจ้าของเดียวกัน หลายสัตว์)
    """
    try:
        import appointment_db as adb
        all_customers = adb.get_all_customers()
        pending = [c for c in all_customers
                   if c.get("hn") and not (c.get("owner_name") or c.get("pet_name"))]
        filled = 0
        if pending:
            print(f"   🔄 Backfill: {len(pending)} customers ที่ยังไม่มีชื่อ...")
            for cust in pending:
                hn = cust["hn"]
                info = adb.verify_hn_with_drx(hn)
                if not info or not info.get("owner_name"):
                    continue
                adb.register_customer(
                    line_user_id=cust["line_user_id"],
                    hn=hn,
                    owner_name=info.get("owner_name", ""),
                    pet_name=info.get("pet_name", ""),
                )
                print(f"      ✅ HN {hn} → {info.get('pet_name','')} / {info.get('owner_name','')}")
                filled += 1
            if filled:
                print(f"   ✅ Backfill สำเร็จ {filled} ราย")
                all_customers = adb.get_all_customers()   # reload

        # ── Auto-register sibling HNs (เจ้าของมีหลายสัตว์) ──
        n = _auto_register_sibling_hns(adb, all_customers)
        if n:
            print(f"   🔗 Auto-register siblings: {n} HNs ใหม่")
    except Exception as e:
        print(f"   [backfill] error: {e}")


def verify_all_line_uids():
    """ตรวจ LINE UID ของลูกค้าทุกรายว่ายังใช้งานได้อยู่ไหม
    - ถ้า block/ลบบัญชี → เพิ่ม 'LINE_BLOCKED' ใน note
    - ถ้า valid อีกครั้ง → ล้าง LINE_BLOCKED ออก
    - ถ้าไม่มี token หรือ network error → ข้าม (ไม่ตัดสิน)
    """
    try:
        import os
        import appointment_db as adb
        token = (os.environ.get("LINE_OA_TOKEN") or os.environ.get("LINE_TOKEN", "")).strip()
        if not token:
            print("   [verify-uid] ⚠️ ไม่มี LINE_OA_TOKEN — ข้ามการตรวจ UID")
            return

        all_customers = adb.get_all_customers()
        if not all_customers:
            return
        print(f"   🔍 Verify LINE UIDs: {len(all_customers)} customers...")
        blocked_new = 0
        unblocked   = 0
        skip_err    = 0

        for cust in all_customers:
            uid  = cust.get("line_user_id")
            note = str(cust.get("note") or "")
            if not uid:
                continue

            is_valid = adb.verify_line_uid(uid, token)
            if is_valid is None:
                skip_err += 1
                continue   # network error / rate limit — ไม่ตัดสิน

            if is_valid is False:
                if "LINE_BLOCKED" not in note:
                    new_note = (note + " LINE_BLOCKED").strip() if note else "LINE_BLOCKED"
                    adb.update_customer_note(uid, new_note)
                    print(f"      🚫 UID {uid[:10]}... HN={cust.get('hn')} → LINE_BLOCKED")
                    adb.log_event("LineBlock", cust.get("hn",""), uid, "user block/ลบบัญชี", "WARN")
                    blocked_new += 1
            else:
                # valid — ถ้าเคย blocked ให้ล้างออก
                if "LINE_BLOCKED" in note:
                    new_note = note.replace("LINE_BLOCKED", "").strip()
                    adb.update_customer_note(uid, new_note)
                    print(f"      ✅ UID {uid[:10]}... HN={cust.get('hn')} → unblocked")
                    adb.log_event("LineUnblock", cust.get("hn",""), uid, "valid อีกครั้ง", "OK")
                    unblocked += 1

        parts = []
        if blocked_new: parts.append(f"blocked ใหม่ {blocked_new} ราย")
        if unblocked:   parts.append(f"unblocked {unblocked} ราย")
        if skip_err:    parts.append(f"ข้าม {skip_err} (network)")
        print(f"   [verify-uid] {'เสร็จ: ' + ', '.join(parts) if parts else 'ไม่มีการเปลี่ยนแปลง'}")
    except Exception as e:
        print(f"   [verify-uid] error: {e}")


def run_drx_sync(fetch_fresh: bool = True):
    """ดึง DRX → DRX_Appointments + mirror ไป Google Sheet"""
    print("─── DRX SYNC ───")
    if fetch_fresh:
        fetch_drx_if_needed(force=True)
    drx = load_drx_appointments()
    print(f"   loaded {len(drx)} appointments จาก DRX")
    write_drx_sheet(drx)
    _mirror_drx_to_gsheet()
    # backfill ชื่อลูกค้าที่ลงทะเบียนแต่ยังไม่มีข้อมูลใน DRX
    backfill_customer_names()
    # ตรวจ LINE UID ของลูกค้าทุกราย (block/ลบบัญชี → mark LINE_BLOCKED)
    verify_all_line_uids()
    try:
        import appointment_db as adb
        adb.log_event("DRX_Sync", "", "", f"DRX:{len(drx)}", "OK")
    except Exception:
        pass


def run_queue_build():
    """Build Send_Queue จาก DRX_Appointments + mirror ไป Google Sheet
    บน Railway: DRX_Appointments มาจาก Google Sheet (PC sync ก่อน 20:15)
    """
    print("─── QUEUE BUILD (T+1 catch-up + T+2) ───")
    # 0a. sync customers จาก Google Sheet
    gs_synced = sync_customers_from_gsheet()
    if gs_synced == 0:
        sync_customers_from_railway()
    # 0b. pull DRX จาก Google Sheet (Railway เข้า DRX ไม่ได้)
    sync_drx_from_gsheet()
    # 1. build queue
    added, total = collect_send_queue()
    print(f"   Send_Queue: +{added} new (จาก {total} candidates ใน T+1/T+2)")
    _mirror_queue_to_gsheet()
    try:
        import appointment_db as adb
        adb.log_event("Queue_Build", "", "", f"Queue+{added}", "OK")
    except Exception:
        pass


def main():
    args = set(sys.argv[1:])
    drx_only   = "--drx-only" in args
    queue_only = "--queue-only" in args
    fetch_fresh = "--fetch" in args or drx_only   # drx-only เสมอเอา fresh

    print("=" * 55)
    print("  📅 Appointment Sync")
    print("=" * 55)

    if not XLSX.exists():
        print(f"❌ {XLSX.name} missing — run init_appointments_xlsx.py first")
        sys.exit(1)

    if drx_only:
        run_drx_sync(fetch_fresh=True)
    elif queue_only:
        run_queue_build()
    else:
        # full sync (manual run / first time)
        run_drx_sync(fetch_fresh=fetch_fresh)
        run_queue_build()

    print()
    print("[DONE] Sync completed")


if __name__ == "__main__":
    main()

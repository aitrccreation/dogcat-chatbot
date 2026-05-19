"""
สร้าง appointments.xlsx เริ่มต้น — 5 sheets
รัน 1 ครั้งเท่านั้น (ไม่ทับไฟล์เดิม)
"""
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

FILE = Path(__file__).parent / "appointments.xlsx"

if FILE.exists():
    print(f"[SKIP] {FILE.name} มีอยู่แล้ว — ไม่ overwrite")
    print("       ลบไฟล์เดิมก่อนถ้าต้องการสร้างใหม่")
    raise SystemExit(0)

wb = Workbook()
wb.remove(wb.active)  # ลบ Sheet default

# ── Styles ─────────────────────────────────────
HEADER_FONT = Font(name="TH Sarabun New", size=14, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2563EB")   # blue-600
CELL_FONT   = Font(name="TH Sarabun New", size=12)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def add_sheet(name: str, headers: list[tuple[str, int]]):
    """headers = [(col_name, width), ...]"""
    ws = wb.create_sheet(name)
    for col_idx, (col_name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    return ws


# ── Sheet 1: Customers (mapping) ─────────────────
ws_c = add_sheet("Customers", [
    ("line_user_id",    36),
    ("hn",              16),
    ("owner_name",      24),
    ("pet_name",        16),
    ("pet_type",        10),
    ("phone",           14),
    ("registered_at",   20),
    ("last_active",     20),
    ("note",            30),
])


# ── Sheet 2: DRX_Appointments (auto-sync จาก DRX) ─────────────────
ws_d = add_sheet("DRX_Appointments", [
    ("appt_id",         12),
    ("hn",              16),
    ("owner_name",      24),
    ("pet_name",        16),
    ("vet",             20),
    ("service",         30),
    ("appt_date",       14),
    ("appt_time",       10),
    ("source",          10),
    ("synced_at",       20),
])


# ── Sheet 3: Manual (หมอเพิ่มเอง) ─────────────────
ws_m = add_sheet("Manual", [
    ("appt_id",         12),
    ("hn",              16),
    ("owner_name",      24),
    ("pet_name",        16),
    ("vet",             20),
    ("service",         30),
    ("appt_date",       14),
    ("appt_time",       10),
    ("notes",           30),
    ("added_by",        16),
    ("added_at",        20),
])

# Sample row guide สำหรับหมอ
ws_m.cell(row=2, column=1, value="(ตัวอย่าง)")
ws_m.cell(row=2, column=2, value="690200-1")
ws_m.cell(row=2, column=3, value="คุณสมชาย ทดสอบ")
ws_m.cell(row=2, column=4, value="ขนมปัง")
ws_m.cell(row=2, column=5, value="หมอนารีรัตน์")
ws_m.cell(row=2, column=6, value="ตรวจวัคซีน")
ws_m.cell(row=2, column=7, value="2026-05-21")
ws_m.cell(row=2, column=8, value="10:00")
ws_m.cell(row=2, column=9, value="ลูกค้าโทรนัด")
ws_m.cell(row=2, column=10, value="หมอนารีรัตน์")
for col in range(1, 12):
    ws_m.cell(row=2, column=col).font = Font(name="TH Sarabun New", size=11, italic=True, color="9CA3AF")


# ── Sheet 4: Send_Queue (รวมจาก DRX + Manual + status) ─────────────────
ws_q = add_sheet("Send_Queue", [
    ("queue_id",        12),
    ("appt_date",       14),
    ("appt_time",       10),
    ("hn",              16),
    ("owner_name",      24),
    ("pet_name",        16),
    ("vet",             20),
    ("service",         30),
    ("line_user_id",    36),
    ("status",          14),     # Pending / Sent / Confirmed / Reschedule / NoLine
    ("sent_round_1_at", 20),     # T-3
    ("sent_round_2_at", 20),     # T-1
    ("sent_round_3_at", 20),     # T-0 (ถ้ายังไม่ตอบ)
    ("response_at",     20),
    ("response",        14),
    ("source",          10),     # drx / manual
])

# Conditional formatting สำหรับ status column
from openpyxl.formatting.rule import FormulaRule
status_col = "J"   # column J = status
ws_q.conditional_formatting.add(
    f"{status_col}2:{status_col}1000",
    CellIsRule(operator="equal", formula=['"Confirmed"'],
               fill=PatternFill("solid", fgColor="86EFAC"))   # green
)
ws_q.conditional_formatting.add(
    f"{status_col}2:{status_col}1000",
    CellIsRule(operator="equal", formula=['"Reschedule"'],
               fill=PatternFill("solid", fgColor="FCD34D"))   # amber
)
ws_q.conditional_formatting.add(
    f"{status_col}2:{status_col}1000",
    CellIsRule(operator="equal", formula=['"NoLine"'],
               fill=PatternFill("solid", fgColor="FCA5A5"))   # red
)


# ── Sheet 5: Sent_Log (audit) ─────────────────
ws_l = add_sheet("Sent_Log", [
    ("timestamp",       20),
    ("event",           20),     # Sync / Send / Response / Error
    ("hn",              16),
    ("line_user_id",    36),
    ("detail",          50),
    ("result",          12),
])


# บันทึก
FILE.parent.mkdir(parents=True, exist_ok=True)
wb.save(FILE)
print(f"[OK] สร้าง {FILE}")
print(f"     {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
print()
print("📋 โครงสร้าง:")
print("  1. Customers         — mapping userId ↔ HN (Phase A)")
print("  2. DRX_Appointments  — auto-sync จาก DRX")
print("  3. Manual            — หมอเพิ่มเอง (เปิด Excel แก้)")
print("  4. Send_Queue        — รายการที่จะส่ง LINE")
print("  5. Sent_Log          — audit log")

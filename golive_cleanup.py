"""
Go-Live Cleanup Script
======================
1. ลบ test_register_user จาก xlsx + Google Sheet
2. ล้าง Send_Queue rows ที่ appt_date < today
3. ล้าง Sent_Log (test logs)
4. Sync ทุกอย่างไป Google Sheet ใหม่
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import os
from datetime import date, datetime
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent
XLSX = BASE / "appointments.xlsx"

os.chdir(BASE)
from dotenv import load_dotenv
load_dotenv()

TODAY = date.today().isoformat()
print(f"[Go-Live Cleanup] วันที่ {TODAY}")
print("=" * 60)

# ─────────────────────────────────────────────
# STEP 1: ลบ test_register_user จาก xlsx
# ─────────────────────────────────────────────
print("\n[1] ลบ test_register_user จาก xlsx Customers...")
wb = openpyxl.load_workbook(XLSX)
ws_cust = wb["Customers"]

# หา row ที่ line_user_id == 'test_register_user'
rows_to_delete = []
for i, row in enumerate(ws_cust.iter_rows(min_row=2, values_only=True), start=2):
    uid = str(row[0] or "")
    if uid == "test_register_user":
        rows_to_delete.append(i)

if rows_to_delete:
    for row_idx in reversed(rows_to_delete):  # ลบจากล่างขึ้นบน
        ws_cust.delete_rows(row_idx)
    print(f"   ✅ ลบ {len(rows_to_delete)} test row(s) ออกจาก Customers")
else:
    print("   ℹ️  ไม่มี test_register_user อยู่แล้ว")

# แสดง customers ที่เหลือ
print("   Customers ที่เหลือ:")
for row in ws_cust.iter_rows(min_row=2, values_only=True):
    if row[0]:
        uid_short = str(row[0])[:20] + "..."
        print(f"     - {uid_short}  HN={row[1]}")

# ─────────────────────────────────────────────
# STEP 2: ล้าง Send_Queue rows เก่า (appt_date < today)
# ─────────────────────────────────────────────
print("\n[2] ล้าง Send_Queue เก่า (appt_date < วันนี้)...")
ws_queue = wb["Send_Queue"]
headers = [c.value for c in ws_queue[1]]
appt_date_col = headers.index("appt_date") if "appt_date" in headers else 1

queue_rows_before = ws_queue.max_row - 1
rows_to_keep = []   # เก็บ rows ที่ appt_date >= today
rows_deleted = 0

for row in ws_queue.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    appt_date = str(row[appt_date_col] or "")
    if appt_date >= TODAY:
        rows_to_keep.append(row)
    else:
        rows_deleted += 1

# ล้าง sheet แล้วเขียนใหม่
ws_queue.delete_rows(2, ws_queue.max_row)
for r in rows_to_keep:
    ws_queue.append(r)

print(f"   ✅ ลบ {rows_deleted} rows เก่า | เหลือ {len(rows_to_keep)} rows (วันนี้+อนาคต)")

# ─────────────────────────────────────────────
# STEP 3: ล้าง Sent_Log (fresh start)
# ─────────────────────────────────────────────
print("\n[3] Reset Sent_Log...")
ws_log = wb["Sent_Log"]
log_count = ws_log.max_row - 1
ws_log.delete_rows(2, ws_log.max_row)
# เพิ่ม 1 row บันทึก go-live event
ws_log.append([
    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "GoLive", "", "", "System cleanup — production start", "OK"
])
print(f"   ✅ ลบ {log_count} test log rows | เพิ่ม GoLive marker")

# บันทึก xlsx
wb.save(XLSX)
print("\n✅ xlsx saved")

# ─────────────────────────────────────────────
# STEP 4: Sync ไป Google Sheet
# ─────────────────────────────────────────────
print("\n[4] Sync ไป Google Sheet...")
try:
    import gsheet_db
    if not gsheet_db.is_enabled():
        print("   ⚠️  Google Sheet ไม่ได้ต่อ — ข้าม sync")
    else:
        # 4a: ลบ test_register_user จาก Google Sheet Customers
        print("   [4a] ลบ test_register_user จาก Google Sheet...")
        sheet = gsheet_db._get_sheet()
        if sheet:
            all_rows = sheet.get_all_values()  # รวม header
            # หา index ของ rows ที่ต้องลบ (นับจาก 1, row 1 = header)
            to_delete_gsheet = []
            for i, row in enumerate(all_rows[1:], start=2):  # start=2 เพราะ row 1 = header
                if row and str(row[0]) == "test_register_user":
                    to_delete_gsheet.append(i)

            if to_delete_gsheet:
                for row_idx in reversed(to_delete_gsheet):
                    sheet.delete_rows(row_idx)
                print(f"      ✅ ลบ {len(to_delete_gsheet)} test row(s) จาก Google Sheet Customers")
            else:
                print("      ℹ️  ไม่มี test_register_user ใน Google Sheet แล้ว")

        # 4b: Sync Send_Queue ไป Google Sheet
        print("   [4b] Sync Send_Queue...")
        SEND_QUEUE_HEADERS = [
            "queue_id", "appt_date", "appt_time", "hn", "owner_name", "pet_name",
            "vet", "service", "line_user_id", "status",
            "sent_round_1_at", "sent_round_2_at", "sent_round_3_at",
            "response_at", "response", "source",
        ]
        rows_to_sync = [list(r) for r in rows_to_keep]
        ok = gsheet_db.sync_send_queue(rows_to_sync)
        print(f"      {'✅' if ok else '❌'} sync_send_queue: {len(rows_to_sync)} rows")

        # 4c: บันทึก GoLive event ใน Log
        print("   [4c] บันทึก GoLive event...")
        gsheet_db.log_event("GoLive", detail="System cleanup — production start", result="OK")
        print("      ✅ logged")

        # 4d: ลบ test log entries จาก Google Sheet Log (optional — skip เพื่อเก็บ history)
        print("   [4d] Google Sheet Log: เก็บ history ไว้ (ไม่ล้าง)")

except Exception as e:
    print(f"   ❌ Google Sheet sync error: {e}")

# ─────────────────────────────────────────────
# STEP 5: แสดงสถานะสุดท้าย
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("📋 สรุป Go-Live Cleanup:")
wb2 = openpyxl.load_workbook(XLSX)

ws = wb2["Customers"]
cust_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[0])
print(f"  Customers    : {cust_count} accounts (ลูกค้าจริง)")

ws = wb2["Send_Queue"]
q_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(r))
print(f"  Send_Queue   : {q_count} rows (วันนี้+อนาคต)")

ws = wb2["Sent_Log"]
l_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(r))
print(f"  Sent_Log     : {l_count} row (GoLive marker)")

ws = wb2["DRX_Appointments"]
drx_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(r))
print(f"  DRX_Appts    : {drx_count} rows (ข้อมูล DRX ล่าสุด)")

print("\n🚀 ระบบพร้อม Go-Live แล้วครับ!")

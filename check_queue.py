import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import load_workbook
from datetime import date

wb = load_workbook("appointments.xlsx")
ws = wb["Send_Queue"]
headers = [c.value for c in ws[1]]

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0]:
        rows.append(dict(zip(headers, r)))

today = date.today()
pending = [r for r in rows if str(r.get("status") or "").lower() not in ("sent","skip","cancel","cancelled")]
sent    = [r for r in rows if str(r.get("status") or "").lower() == "sent"]
noline  = [r for r in pending if str(r.get("status") or "").lower() == "noline"]
active  = [r for r in pending if str(r.get("status") or "").lower() not in ("noline","reschedule")]

print(f"วันนี้: {today}  |  Total: {len(rows)}  |  Sent: {len(sent)}  |  Pending: {len(pending)}")
print(f"NoLine: {len(noline)} (ยังไม่มีเบอร์ LINE)  |  Active (มี LINE): {len(active)}")
print()
print("=== Active Pending (มี LINE) ===")
for r in active:
    appt = str(r.get("appt_date",""))[:10]
    lid  = str(r.get("line_user_id") or "")
    print(f"  qid={str(r['queue_id']):6}  HN={r['hn']}  {str(r['pet_name']):<14}  นัด={appt}  status=[{r['status']}]  LINE={'✅' if lid else '❌'}")
print()
print("=== Sent ===")
for r in sent:
    print(f"  qid={str(r['queue_id']):6}  HN={r['hn']}  {str(r['pet_name']):<14}  ส่ง={str(r.get('sent_round_1_at',''))[:16]}")
print()
print("=== NoLine (ยังไม่มี LINE ID — รอ match) ===")
by_date = {}
for r in noline:
    d = str(r.get("appt_date",""))[:10]
    by_date.setdefault(d, []).append(r)
for d in sorted(by_date):
    print(f"  {d}: {len(by_date[d])} ราย")

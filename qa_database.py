"""
Q&A Database — โหลดจาก fb_replies_export.xlsx Sheet1
  - Column A = ชื่อรูปที่ต้องส่งตาม (ตรงกับไฟล์ใน static/images/)
  - Column C = คำถาม
  - Column D = คำตอบ
"""
import re
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(__file__).parent / "fb_replies_export.xlsx"


# ──────────────────────────────────────────
#  IMAGE_MAP — ชื่อในคอลัมน์ A → path ของรูป
#  (ชื่อต้องตรงกับ Column A ใน Excel)
# ──────────────────────────────────────────
IMAGE_MAP: dict[str, list[str]] = {
    "แพคเกจทำหมันแมว": [
        "/static/images/แพคเกจทำหมันแมว.png",
    ],
    "แพคเกจทำหมันหมาแมว": [
        "/static/images/แพคเกจทำหมันหมาแมว.png",
    ],
    "แพคเกจขูดหินปูน": [
        "/static/images/แพคเกจขูดหินปูน.jpg",
    ],
    "แพคเกจวัคซีน": [
        "/static/images/แพคเกจวัคซีน.jpg",
    ],
    "แพคเกจตรวจสุขภาพ1 และ 2": [
        "/static/images/แพคเกจตรวจสุขภาพ1.jpg",
        "/static/images/แพคเกจตรวจสุขภาพ2.jpg",
    ],
    "อาบน้ำตัดขน": [
        "/static/images/อาบน้ำตัดขน.jpg",
    ],
    "บริการของโรงพยาบาล": [
        "/static/images/บริการของโรงพยาบาล.jpg",
    ],
}


# ──────────────────────────────────────────
#  Short labels — แสดงในเมนูแทนคำถามเต็ม
#  key = row number ใน Excel (= id ของ QA item)
#  Row 2 = id 2 = No.1 ใน Excel
# ──────────────────────────────────────────
SHORT_LABELS: dict[int, str] = {
    2:  "ทำหมันแมว (ราคาทั่วไป)",
    3:  "วิธีทำหมัน / ดมยาสลบ",
    4:  "วัคซีนลูกสุนัข (ปีแรก)",
    5:  "บริการฝากเลี้ยง",
    6:  "ทำหมันแมวเพศผู้ + ขูดหินปูน",
    7:  "ผ่าคลอด (ราคา)",
    8:  "ผ่าคลอดแมว (รายการ+ราคา)",
    9:  "วัคซีนไม่ครบ ทำหมันได้ไหม",
    10: "น้องป่วย ควรพามาเลยไหม",
    11: "ราคาทำหมัน (รวมตรวจเลือดไหม)",
    12: "ทำหมันแมว อายุ + การเตรียมตัว",
    13: "ทำหมันแมวเพศเมีย (ฝากข้ามคืน)",
    14: "ทำหมันแมวเพศเมีย (รับกลับบ้าน)",
    15: "ที่ตั้งโรงพยาบาล",
    16: "ทำหมันแมวเพศผู้ (ฝากข้ามคืน)",
    17: "บริการขูดหินปูน",
    18: "ทำหมันแมวเพศผู้ (รับกลับบ้าน)",
    19: "อาบน้ำสุนัข เงื่อนไข + เวลา",
    20: "ขูดหินปูนสุนัข (<10 กก.)",
    21: "การเตรียมตัวก่อนผ่าตัด",
    22: "สวนท่อปัสสาวะ (ราคา)",
    23: "วัคซีนแมว (รายการ+ราคา)",
    24: "แพ็คเกจวัคซีน",
    25: "บริการผ่าตัดกระดูก",
    26: "วิธีใช้และเก็บยาหยด",
    27: "เอกซเรย์/อุลตราซาวด์ดูท้อง",
    28: "วัคซีนเอดส์แมว และ FIP",
    29: "ทำหมันสุนัข 30-40 กก.",
    30: "ยาหยดกำจัดเห็บหมัดแมว",
    31: "วัคซีนครั้งแรก (เพิ่งรับน้องมา)",
    32: "เม่นแคระเป็นไร (รักษา)",
    33: "ยาหยดเห็บหมัดเม่นแคระ",
    34: "โปรแกรมตรวจสุขภาพ",
    35: "แมวเป็นสัด ทำหมันได้ไหม",
    36: "ตรวจฮอร์โมนผสมพันธุ์สุนัข",
    37: "ผ่าคลอดสุนัข + อุลตราซาวด์",
    38: "ยากิน กำจัดเห็บหมัดสุนัข",
    39: "ตรวจเลือด CBC / ตับ-ไต / เอดส์แมว",
    40: "อุลตราซาวด์ตรวจท้อง",
    41: "ยาหยด กำจัดเห็บหมัดสุนัข",
    42: "ยาคุม แมว/สุนัข (ปลอดภัยไหม)",
    43: "ค่าตรวจร่างกาย (ราคาเริ่มต้น)",
    44: "ยาหยดหมัดไร + ยาถ่ายพยาธิ",
    45: "ทำหมันสุนัข 20-30 กก.",
    46: "ทำหมันสุนัข (<10 กก.)",
    47: "ทำหมันสุนัข 10-20 กก.",
    48: "ฝากเลี้ยงแมว (ราคา/คืน)",
    49: "จองคิวทำหมัน",
    50: "จองคิวอาบน้ำ-ตัดขน",
}


# ──────────────────────────────────────────
#  โหลด Q&A จาก Excel
#  Column A = ชื่อรูป | Column C = คำถาม | Column D = คำตอบ
# ──────────────────────────────────────────
def load_qa_list():
    """โหลด Q&A — return list of dicts"""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Sheet1"]
    qa_list = []
    for r in range(2, ws.max_row + 1):
        img_name = (ws.cell(r, 1).value or "").strip()
        q        = (ws.cell(r, 3).value or "").strip()
        a        = (ws.cell(r, 4).value or "").strip()
        if not q or not a:
            continue
        qa_list.append({
            "id":       r,
            "question": q,
            "answer":   a,
            "images":   IMAGE_MAP.get(img_name, []),   # ← จาก Column A โดยตรง
            "label":    SHORT_LABELS.get(r, q),        # ← label สั้นสำหรับเมนู
        })
    return qa_list


QA_LIST = load_qa_list()


# ──────────────────────────────────────────
#  Keyword search — หา candidate questions
# ──────────────────────────────────────────
# Synonym map — ขยาย keyword ของ user เพื่อ match ได้กว้างขึ้น
SYNONYMS = {
    "หมา":     ["หมา", "สุนัข", "น้องหมา"],
    "สุนัข":   ["หมา", "สุนัข", "น้องหมา"],
    "เห็บ":    ["เห็บ", "หมัด", "ไร"],
    "หมัด":    ["เห็บ", "หมัด", "ไร"],
    "ราคา":    ["ราคา", "ค่า", "เท่าไหร่", "เท่าไร", "กี่บาท"],
    "ทำหมัน":  ["ทำหมัน", "ผ่าหมัน", "หมัน"],
    "ผ่าคลอด": ["ผ่าคลอด", "คลอด"],
    "วัคซีน":  ["วัคซีน", "ฉีดยา"],
    "อาบน้ำ":  ["อาบน้ำ", "ตัดขน"],
    "ฝาก":     ["ฝาก", "ฝากเลี้ยง", "พักค้าง"],
    "จอง":     ["จอง", "นัด", "คิว"],
}


def normalize(text: str) -> str:
    """lowercase + remove punctuation"""
    text = text.lower()
    text = re.sub(r"[^\w\s฀-๿]", " ", text)
    return text


def extract_keywords(text: str) -> list:
    """ดึง keyword จาก user message (รวม synonym expansion)"""
    t = normalize(text)
    keywords = set()
    # คำ Thai 2+ chars
    for w in re.findall(r"[฀-๿]{2,}", t):
        keywords.add(w)
        for canonical, syns in SYNONYMS.items():
            if w in syns:
                keywords.update(syns)
    # คำ English
    for w in re.findall(r"[a-z]{3,}", t):
        keywords.add(w)
    return list(keywords)


def score_qa(qa: dict, keywords: list) -> int:
    """score คำถามตามจำนวน keyword ที่ตรง"""
    q_norm = normalize(qa["question"])
    score = 0
    for kw in keywords:
        if kw in q_norm:
            score += 10
            if q_norm.startswith(kw):
                score += 5
    return score


def find_candidates(user_msg: str, top_k: int = 4) -> list:
    """หา candidate questions ที่ใกล้เคียงกับ user message"""
    keywords = extract_keywords(user_msg)
    if not keywords:
        return []
    scored = []
    for qa in QA_LIST:
        s = score_qa(qa, keywords)
        if s > 0:
            scored.append((s, qa))
    scored.sort(key=lambda x: -x[0])
    return [qa for _, qa in scored[:top_k]]


def find_exact(user_msg: str) -> dict | None:
    """หา QA ที่คำถามตรงเป๊ะ (normalized) — ใช้ skip AI สำหรับปุ่ม/quick reply/copy-paste
    คืน QA เฉพาะเมื่อ normalized user == normalized question ทั้งหมด (ปลอดภัย 100%
    ไม่มีความเสี่ยงตอบผิด weight/เพศ/ประเภท เพราะตรงตัวอักษรเป๊ะ)
    """
    un = normalize(user_msg).strip()
    un = re.sub(r"\s+", " ", un)
    if len(un) < 4:
        return None
    for qa in QA_LIST:
        qn = re.sub(r"\s+", " ", normalize(qa["question"]).strip())
        if qn and un == qn:
            return qa
    return None


def find_by_id(qa_id: int) -> dict | None:
    """หา QA ตาม id"""
    for qa in QA_LIST:
        if qa["id"] == qa_id:
            return qa
    return None


# ──────────────────────────────────────────
#  Affirmative / Negative detection
# ──────────────────────────────────────────
AFFIRMATIVE = [
    r"^ใช่", r"^ถูก", r"^ค่ะ$", r"^ครับ$", r"^ใช่ค่ะ", r"^ใช่ครับ",
    r"^ok$", r"^okay$", r"^yes$", r"^y$", r"^correct",
    r"ถูกต้อง", r"ใช่แล้ว",
]

NEGATIVE = [
    r"^ไม่", r"^ไม่ใช่", r"^ผิด", r"^no$", r"^n$", r"^nope",
    r"ไม่ตรง", r"ไม่ถูก",
]


def is_affirmative(text: str) -> bool:
    t = text.strip().lower()
    return any(re.search(p, t) for p in AFFIRMATIVE)


def is_negative(text: str) -> bool:
    t = text.strip().lower()
    return any(re.search(p, t) for p in NEGATIVE)


def parse_choice_number(text: str) -> int | None:
    """ดึงเลขจาก user input '1', 'ข้อ 1', '1.', '#1'"""
    t = text.strip()
    m = re.search(r"\b([1-9])\b", t)
    if m:
        return int(m.group(1))
    m = re.search(r"(ข้อ|ตัวเลือก|choice|option)\s*([1-9])", t.lower())
    if m:
        return int(m.group(2))
    return None


if __name__ == "__main__":
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    print(f"Loaded {len(QA_LIST)} Q&A pairs")
    print()
    print("=== Items with images ===")
    for q in QA_LIST:
        if q["images"]:
            print(f"  [{q['id']:2d}] {q['label']}")
            for img in q["images"]:
                print(f"        📷 {img}")
    print()
    for test in ["ทำหมันแมว", "ทำหมันสุนัข", "วัคซีน", "ผ่าคลอด", "จองคิว"]:
        print(f"=== Query: '{test}' ===")
        cands = find_candidates(test)
        for c in cands:
            print(f"  [{c['id']:2d}] {c['label']}")
        print()
